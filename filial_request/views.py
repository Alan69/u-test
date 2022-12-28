from django.shortcuts import render
from .forms import RequestForm
from django.http import HttpResponse
from django.contrib.auth.models import User
from userprofile.models import Profile
from openpyxl import load_workbook

# Create your views here.
def request_page(request):
    if request.method=='POST':
        form = RequestForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
        return HttpResponse("Запрос отправлен")
    else:
        form = RequestForm()
        return render(request,'filial/request.html', {'form':form})


def add_students(request):
    if request.method == 'POST':
        workbook = load_workbook(request.FILES['document'])
        worksheet = workbook.active

        student_region = request.user.profile.region
        school_region = request.user.profile.school

        for row in worksheet.iter_rows():
            username = row[0].value
            password = row[0].value
            first_name = row[1].value
            last_name = row[2].value
            class_name = row[3].value

            if User.objects.filter(username=username).exists():
                continue

            user = User.objects.create_user(
                username=username,
                password=password,
                first_name=first_name,
                last_name=last_name
            )
            user.save()

            # if Profile.objects.filter(user=user).exists():
            #     continue

            # profile = Profile(user=user, class_name=class_name, region=student_region, school=school_region)
            # profile.save()

        return HttpResponse("Ученики добавлены")
    else:
        return render(request,'filial/addstudent.html')